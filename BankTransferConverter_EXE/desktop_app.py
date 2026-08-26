import os
import re
import threading
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

APP_TITLE = "Holyoke Bank Transfer Converter"

DEFAULT_PAIR_SCRUBS = {
    frozenset(["1171", "1220"]),
    frozenset(["1171", "1270"]),
}
DEFAULT_GROUP_SCRUB = {"7189", "4237", "4245"}

EMBEDDED_BANK_CODES = {
    "1063": "SCHMER",
    "1908": "PEOP2",
    "2441": "EMP99",
    "2882": "PGECD",
    "4010": "COHMER",
    "4503": "STAB5",
    "4531": "UNICD2",
    "4892": "EMP03",
    "4909": "EMP05",
    "4917": "EMP06",
    "5123": "CDBG1",
    "5149": "HOME1",
    "5628": "TAIL",
    "5755": "CAPSTAB",
    "6049": "HGEMER",
    "6125": "WWTP1",
    "6223": "CPA",
    "6600": "TT",
    "6754": "UNI2",
    "6885": "HGEDEP",
    "7067": "SCHPR",
    "7075": "CITYPR",
    "7083": "HGEPR",
    "7355": "STUD2",
    "7428": "PEOP",
    "7805": "EXCU",
    "7821": "WWTP",
    "7839": "SLUN",
    "8035": "STUD1",
    "8128": "ESB2",
    "8166": "PGEAP",
    "8207": "MJFEE",
    "8318": "SPROP",
    "8464": "HEALTH&DEN",
    "8506": "ARPA",
    "8542": "PRVEN",
    "8555": "AP",
    "8571": "PR",
    "8636": "SAP",
    "8652": "CAP",
    "8801": "UNICD",
    "8917": "APSWP",
    "8941": "DEPSWP",
    "8989": "ESBCD"
}


def digits_only(value):
    return re.sub(r"\D", "", str(value or ""))


def last4(value):
    d = digits_only(value)
    return d[-4:] if len(d) >= 4 else str(value).strip()


def normalize_amount(series):
    return (
        series.astype(str)
        .str.replace(r"[^0-9\.\-]", "", regex=True)
        .replace("", pd.NA)
        .astype(float)
    )


def parse_extra_pairs(text):
    pairs = set()
    for line in text.splitlines():
        parts = [p.strip() for p in line.split(",") if p.strip()]
        if len(parts) == 2:
            pairs.add(frozenset(parts))
    return pairs


def should_scrub(from_code, to_code, from4, to4, extra_pairs):
    pair_codes = frozenset([str(from_code), str(to_code)])
    pair_nums = frozenset([str(from4), str(to4)])

    if pair_nums in DEFAULT_PAIR_SCRUBS:
        return True
    if str(from4) in DEFAULT_GROUP_SCRUB and str(to4) in DEFAULT_GROUP_SCRUB:
        return True
    if pair_codes in extra_pairs or pair_nums in extra_pairs:
        return True
    return False


def convert_file(raw_csv, output_path, extra_scrub_text=""):
    mapping = EMBEDDED_BANK_CODES
    extra_pairs = parse_extra_pairs(extra_scrub_text)

    df = pd.read_csv(raw_csv)
    required = {"Will Process On", "Amount", "From Account", "To Account"}
    missing = required - set(df.columns)
    if missing:
        raise ValueError("Raw CSV is missing required columns: " + ", ".join(sorted(missing)))

    df["Will Process On"] = pd.to_datetime(df["Will Process On"], errors="coerce")
    df["Amount"] = normalize_amount(df["Amount"])
    df["From4"] = df["From Account"].apply(last4)
    df["To4"] = df["To Account"].apply(last4)
    df["FromCode"] = df["From4"].apply(lambda a: mapping.get(a, a))
    df["ToCode"] = df["To4"].apply(lambda a: mapping.get(a, a))

    encountered = sorted(set(df["From4"]).union(set(df["To4"])))
    missing_codes = [a for a in encountered if a.isdigit() and len(a) == 4 and a not in mapping]

    records = []
    scrubbed = []

    for _, r in df.iterrows():
        frm4 = str(r["From4"]).strip()
        to4 = str(r["To4"]).strip()
        frm = str(r["FromCode"]).strip()
        to = str(r["ToCode"]).strip()
        amt = abs(float(r["Amount"]))
        dt = r["Will Process On"]
        desc = f"TRANSFER FROM {frm} TO {to}"

        if should_scrub(frm, to, frm4, to4, extra_pairs):
            scrubbed.append({"Date": dt, "Description": desc, "Amount": amt})
            continue

        # TRANSFER FROM = negative, TRANSFER TO = positive
        # Keep the pair adjacent and in debit-then-credit order.
        records.append({"Date": dt, "Description": desc, "Account": frm, "Amount": -amt})
        records.append({"Date": dt, "Description": desc, "Account": to, "Amount": amt})

    out = pd.DataFrame(records)
    if not out.empty:
        out["Month"] = out["Date"].dt.to_period("M")
        out["Sequence"] = range(len(out))
        out = out.sort_values(["Date", "Sequence"], kind="stable")

    wb = Workbook()
    wb.remove(wb.active)

    headers1 = [
        "Date", "Source/Description", "Check #", "Check #", "Receipts",
        "Disbursements", "Transfer", "Deposit", "Deposits", None, None, "Bank"
    ]
    headers2 = [
        None, None, "or Batch #", "or Batch #", None, "(Warrants)",
        None, "date", None, None, None, "Account"
    ]

    if out.empty:
        ws = wb.create_sheet("No Transactions")
        ws.append(headers1)
        ws.append(headers2)
    else:
        for period in sorted(out["Month"].dropna().unique()):
            month_df = out[out["Month"] == period].copy()
            ws = wb.create_sheet(pd.Period(period).strftime("%b %Y"))
            ws.append(headers1)
            ws.append(headers2)

            for _, r in month_df.iterrows():
                ws.append([
                    r["Date"].to_pydatetime() if pd.notna(r["Date"]) else None,
                    r["Description"],
                    None,
                    "TRANSFER",
                    None,
                    None,
                    float(r["Amount"]),
                    None,
                    None,
                    None,
                    float(r["Amount"]),
                    r["Account"],
                ])

            for rr in range(3, ws.max_row + 1):
                ws.cell(rr, 1).number_format = "mm-dd-yyyy"
                amount = ws.cell(rr, 7).value
                if isinstance(amount, (int, float)):
                    color = "FF0000" if amount < 0 else "000000"
                    ws.cell(rr, 7).font = Font(color=color)
                    ws.cell(rr, 11).font = Font(color=color)
                    ws.cell(rr, 7).number_format = '#,##0.00;[Red]-#,##0.00'
                    ws.cell(rr, 11).number_format = '#,##0.00;[Red]-#,##0.00'

            widths = {1: 13, 2: 42, 3: 12, 4: 14, 5: 14, 6: 16, 7: 16, 8: 12, 9: 14, 10: 4, 11: 16, 12: 16}
            for c, width in widths.items():
                ws.column_dimensions[get_column_letter(c)].width = width

            for c in range(1, 13):
                ws.cell(1, c).font = Font(bold=True)
                ws.cell(2, c).font = Font(italic=True)
                ws.cell(1, c).alignment = Alignment(horizontal="center")
                ws.cell(2, c).alignment = Alignment(horizontal="center")

    wb.save(output_path)
    return len(df), len(records), len(scrubbed), missing_codes


class ConverterApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title(APP_TITLE)
        self.geometry("760x510")
        self.minsize(700, 470)

        self.raw_path = tk.StringVar()
        self.output_path = tk.StringVar()
        self.status = tk.StringVar(value="Ready")

        self._build_ui()

    def _build_ui(self):
        pad = {"padx": 12, "pady": 8}

        title = ttk.Label(self, text=APP_TITLE, font=("Segoe UI", 18, "bold"))
        title.pack(pady=(18, 4))
        ttk.Label(self, text="Raw bank CSV -> CB-ready monthly transfer workbook").pack(pady=(0, 14))

        frm = ttk.Frame(self)
        frm.pack(fill="both", expand=True, padx=18)

        ttk.Label(frm, text="1. Raw bank transfer CSV").grid(row=0, column=0, sticky="w", **pad)
        ttk.Entry(frm, textvariable=self.raw_path, width=70).grid(row=1, column=0, sticky="ew", padx=12)
        ttk.Button(frm, text="Browse...", command=self.pick_raw).grid(row=1, column=1, padx=6)

        ttk.Label(
            frm,
            text=f"Bank-code mapping is built in ({len(EMBEDDED_BANK_CODES)} accounts)."
        ).grid(row=2, column=0, columnspan=2, sticky="w", padx=12, pady=(6, 10))

        ttk.Label(frm, text="2. Output workbook").grid(row=3, column=0, sticky="w", **pad)
        ttk.Entry(frm, textvariable=self.output_path, width=70).grid(row=4, column=0, sticky="ew", padx=12)
        ttk.Button(frm, text="Save As...", command=self.pick_output).grid(row=4, column=1, padx=6)

        ttk.Label(frm, text="Additional scrub pairs (optional, one pair per line: 1111,2222)").grid(row=5, column=0, sticky="w", **pad)
        self.extra_scrubs = tk.Text(frm, height=5, width=60)
        self.extra_scrubs.grid(row=6, column=0, columnspan=2, sticky="nsew", padx=12)

        rules = (
            "Built-in rules: FROM = negative, TO = positive; debit first / credit second; "
            "mm-dd-yyyy; column D = TRANSFER; negative red / positive black; monthly tabs; "
            "scrubs 1171<->1220, 1171<->1270, and transfers among 7189/4237/4245."
        )
        ttk.Label(frm, text=rules, wraplength=690).grid(row=7, column=0, columnspan=2, sticky="w", padx=12, pady=12)

        self.convert_btn = ttk.Button(frm, text="Convert", command=self.run_conversion)
        self.convert_btn.grid(row=8, column=0, sticky="w", padx=12, pady=10)
        ttk.Label(frm, textvariable=self.status).grid(row=8, column=0, sticky="e", padx=12)

        frm.columnconfigure(0, weight=1)
        frm.rowconfigure(6, weight=1)

    def pick_raw(self):
        path = filedialog.askopenfilename(filetypes=[("CSV files", "*.csv"), ("All files", "*.*")])
        if path:
            self.raw_path.set(path)
            if not self.output_path.get():
                out = str(Path(path).with_name(Path(path).stem + "_CB_Ready.xlsx"))
                self.output_path.set(out)

    def pick_output(self):
        path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel workbook", "*.xlsx")])
        if path:
            self.output_path.set(path)

    def run_conversion(self):
        raw = self.raw_path.get().strip()
        output = self.output_path.get().strip()

        if not raw or not os.path.exists(raw):
            messagebox.showerror("Missing file", "Please select the raw bank transfer CSV.")
            return
        if not output:
            messagebox.showerror("Missing output", "Please choose where to save the output workbook.")
            return

        self.convert_btn.config(state="disabled")
        self.status.set("Converting...")
        extra = self.extra_scrubs.get("1.0", "end").strip()

        def worker():
            try:
                raw_count, output_rows, scrubbed_count, missing_codes = convert_file(raw, output, extra)
                msg = (
                    f"Conversion complete.\n\nRaw transfers: {raw_count}\n"
                    f"Output rows: {output_rows}\nScrubbed transfers: {scrubbed_count}\n"
                    f"Saved to:\n{output}"
                )
                if missing_codes:
                    msg += "\n\nUnmapped 4-digit accounts (send these to Finance to add to the converter):\n" + ", ".join(missing_codes)
                self.after(0, lambda: messagebox.showinfo("Complete", msg))
                self.after(0, lambda: self.status.set("Complete"))
            except Exception as exc:
                self.after(0, lambda: messagebox.showerror("Conversion error", str(exc)))
                self.after(0, lambda: self.status.set("Error"))
            finally:
                self.after(0, lambda: self.convert_btn.config(state="normal"))

        threading.Thread(target=worker, daemon=True).start()


if __name__ == "__main__":
    app = ConverterApp()
    app.mainloop()
